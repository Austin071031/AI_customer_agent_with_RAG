"""
SQLite Database Service for Excel File Storage.

This module provides the SQLite database service for storing Excel file metadata
and data as required by US-009. It handles structured storage of Excel files
with efficient querying capabilities.
"""

import os
import sqlite3
import logging
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path
import json
import openpyxl
from datetime import datetime

from ..models.excel_models import ExcelDocument, ExcelSheetData, SQLiteDataType, DynamicTableSchema, TableCreationResult


class SQLiteDatabaseError(Exception):
    """Custom exception for SQLite database related errors."""
    
    def __init__(self, message: str, error_type: Optional[str] = None):
        self.message = message
        self.error_type = error_type
        super().__init__(self.message)


class SQLiteDatabaseService:
    """
    Manages SQLite database operations for Excel file storage.
    
    This service handles storing Excel file metadata, sheet data, and provides
    search capabilities for structured Excel data.
    """
    
    def __init__(self, db_path: str = "./excel_database.db"):
        """
        Initialize the SQLite database service.
        
        Args:
            db_path: Path to the SQLite database file
        """
        self.logger = logging.getLogger(__name__)
        self.db_path = db_path
        self._init_database()
        
    def _init_database(self) -> None:
        """
        Initialize SQLite database with required tables.
        
        Creates tables for Excel documents and sheet data if they don't exist.
        """
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # Create Excel documents table
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS excel_documents (
                        id TEXT PRIMARY KEY,
                        file_name TEXT NOT NULL,
                        file_size INTEGER NOT NULL,
                        sheet_names TEXT NOT NULL,
                        metadata TEXT NOT NULL,
                        upload_time TIMESTAMP NOT NULL,
                        storage_type TEXT DEFAULT 'sqlite'
                    )
                """)
                
                # Create Excel sheet data table
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS excel_sheet_data (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        file_id TEXT NOT NULL,
                        sheet_name TEXT NOT NULL,
                        headers TEXT NOT NULL,
                        row_count INTEGER NOT NULL,
                        column_count INTEGER NOT NULL,
                        sample_data TEXT NOT NULL,
                        data_types TEXT NOT NULL,
                        FOREIGN KEY (file_id) REFERENCES excel_documents (id) ON DELETE CASCADE
                    )
                """)
                
                # Create Excel row data table for complete data storage
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS excel_row_data (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        file_id TEXT NOT NULL,
                        sheet_name TEXT NOT NULL,
                        row_index INTEGER NOT NULL,
                        row_data TEXT NOT NULL,
                        FOREIGN KEY (file_id) REFERENCES excel_documents (id) ON DELETE CASCADE
                    )
                """)
                
                # Create indexes for better performance
                cursor.execute("CREATE INDEX IF NOT EXISTS idx_file_id ON excel_sheet_data(file_id)")
                cursor.execute("CREATE INDEX IF NOT EXISTS idx_sheet_name ON excel_sheet_data(sheet_name)")
                cursor.execute("CREATE INDEX IF NOT EXISTS idx_upload_time ON excel_documents(upload_time)")
                
                conn.commit()
                self.logger.info(f"SQLite database initialized at: {self.db_path}")
                
        except sqlite3.Error as e:
            self.logger.error(f"Failed to initialize SQLite database: {str(e)}")
            raise SQLiteDatabaseError(f"Database initialization failed: {str(e)}")

    def _sanitize_for_table_name(self, s: str) -> str:
        """
        Sanitize a string for use in SQLite table names.
        
        Args:
            s: Input string
            
        Returns:
            Sanitized string with only alphanumeric characters and underscores
        """
        return "".join(c for c in s if c.isalnum() or c == '_')

    def _reconstruct_file_id(self, sanitized_uuid: str) -> str:
        """
        Reconstruct the original file ID with hyphens from a sanitized UUID.
        
        Args:
            sanitized_uuid: UUID string without hyphens (32 characters)
            
        Returns:
            File ID in format "excel_{uuid}" with hyphens
        """
        # UUID pattern: 8-4-4-4-12
        if len(sanitized_uuid) == 32:
            return f"excel_{sanitized_uuid[:8]}-{sanitized_uuid[8:12]}-{sanitized_uuid[12:16]}-{sanitized_uuid[16:20]}-{sanitized_uuid[20:]}"
        else:
            # If it's not 32 characters, we cannot reconstruct, so return without hyphens
            return f"excel_{sanitized_uuid}"
            
    def store_excel_file(self, file_path: str, file_name: str, file_size: int,
                        sheet_names: List[str], metadata: Dict[str, Any]) -> str:
        """
        Store Excel file metadata and content in SQLite database.
        
        Args:
            file_path: Path to the Excel file
            file_name: Name of the Excel file
            file_size: Size of the file in bytes
            sheet_names: List of sheet names in the Excel file
            metadata: Additional metadata about the file
            
        Returns:
            File ID of the stored Excel document
            
        Raises:
            SQLiteDatabaseError: If storage operation fails
        """
        try:
            # Create ExcelDocument object for validation
            excel_doc = ExcelDocument(
                file_name=file_name,
                file_size=file_size,
                sheet_names=sheet_names,
                metadata=metadata
            )
            
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # Insert Excel document
                cursor.execute("""
                    INSERT INTO excel_documents 
                    (id, file_name, file_size, sheet_names, metadata, upload_time)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (
                    excel_doc.id,
                    excel_doc.file_name,
                    excel_doc.file_size,
                    json.dumps(excel_doc.sheet_names),
                    json.dumps(excel_doc.metadata),
                    excel_doc.upload_time.isoformat()
                ))
                
                # Process and store sheet data
                self._process_and_store_sheet_data(file_path, excel_doc.id, cursor)
                
                conn.commit()
                self.logger.info(f"Stored Excel file: {file_name} with ID: {excel_doc.id}")
                return excel_doc.id
                
        except sqlite3.Error as e:
            self.logger.error(f"Failed to store Excel file: {str(e)}")
            raise SQLiteDatabaseError(f"Excel file storage failed: {str(e)}")
        except Exception as e:
            self.logger.error(f"Unexpected error storing Excel file: {str(e)}")
            raise SQLiteDatabaseError(f"Unexpected error: {str(e)}")
            
    def _process_and_store_sheet_data(self, file_path: str, file_id: str, cursor: sqlite3.Cursor) -> None:
        """
        Process Excel file and store sheet data in database.
        
        Args:
            file_path: Path to the Excel file
            file_id: ID of the Excel document
            cursor: Database cursor for executing queries
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Extract headers (first row)
                headers = []
                for cell in sheet[1]:
                    headers.append(str(cell.value) if cell.value else f"Column_{cell.column}")
                
                # Count rows and columns
                row_count = sheet.max_row - 1  # Exclude header
                column_count = sheet.max_column
                
                # Extract sample data (first 5 rows after header)
                sample_data = []
                data_types = {}
                
                for row_idx in range(2, min(7, sheet.max_row + 1)):
                    row_data = {}
                    for col_idx, header in enumerate(headers, 1):
                        cell_value = sheet.cell(row=row_idx, column=col_idx).value
                        row_data[header] = cell_value
                        
                        # Infer data type for this column
                        if header not in data_types:
                            data_types[header] = self._infer_data_type(cell_value)
                    
                    sample_data.append(row_data)
                
                # Create ExcelSheetData object for validation
                sheet_data = ExcelSheetData(
                    file_id=file_id,
                    sheet_name=sheet_name,
                    headers=headers,
                    row_count=row_count,
                    column_count=column_count,
                    sample_data=sample_data,
                    data_types=data_types
                )
                
                # Insert sheet data
                cursor.execute("""
                    INSERT INTO excel_sheet_data 
                    (file_id, sheet_name, headers, row_count, column_count, sample_data, data_types)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    sheet_data.file_id,
                    sheet_data.sheet_name,
                    json.dumps(sheet_data.headers),
                    sheet_data.row_count,
                    sheet_data.column_count,
                    json.dumps(sheet_data.sample_data),
                    json.dumps(sheet_data.data_types)
                ))
                
        except Exception as e:
            self.logger.error(f"Failed to process sheet data: {str(e)}")
            raise SQLiteDatabaseError(f"Sheet data processing failed: {str(e)}")
            
    def _infer_data_type(self, value: Any) -> str:
        """
        Infer data type from cell value.
        
        Args:
            value: Cell value from Excel
            
        Returns:
            Inferred data type as string (must match SQLiteDataType enum)
        """
        if value is None:
            return "TEXT"  # Use TEXT for NULL values in SQLite
        elif isinstance(value, bool):
            return "INTEGER"  # SQLite stores booleans as integers (0/1)
        elif isinstance(value, int):
            return "INTEGER"
        elif isinstance(value, float):
            return "REAL"
        elif isinstance(value, str):
            # Check if it's a date string
            try:
                datetime.fromisoformat(value.replace('Z', '+00:00'))
                return "TEXT"  # Dates are stored as TEXT in SQLite
            except (ValueError, TypeError):
                return "TEXT"
        elif isinstance(value, datetime):
            return "TEXT"  # Datetimes are stored as TEXT in SQLite
        else:
            return "TEXT"  # Default to TEXT for unknown types
            
    def get_excel_file(self, file_id: str) -> Optional[ExcelDocument]:
        """
        Retrieve Excel file data from database.
        
        Args:
            file_id: ID of the Excel file to retrieve
            
        Returns:
            ExcelDocument object if found, None otherwise
        """
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                cursor.execute("""
                    SELECT id, file_name, file_size, sheet_names, metadata, upload_time, storage_type
                    FROM excel_documents 
                    WHERE id = ?
                """, (file_id,))
                
                row = cursor.fetchone()
                if not row:
                    return None
                    
                return ExcelDocument(
                    id=row[0],
                    file_name=row[1],
                    file_size=row[2],
                    sheet_names=json.loads(row[3]),
                    metadata=json.loads(row[4]),
                    upload_time=datetime.fromisoformat(row[5]),
                    storage_type=row[6]
                )
                
        except sqlite3.Error as e:
            self.logger.error(f"Failed to retrieve Excel file: {str(e)}")
            raise SQLiteDatabaseError(f"Excel file retrieval failed: {str(e)}")
            
    def list_excel_files(self) -> List[ExcelDocument]:
        """
        List all Excel files in database.
        
        Returns:
            List of ExcelDocument objects
        """
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                cursor.execute("""
                    SELECT id, file_name, file_size, sheet_names, metadata, upload_time, storage_type
                    FROM excel_documents 
                    ORDER BY upload_time DESC
                """)
                
                excel_files = []
                for row in cursor.fetchall():
                    excel_files.append(ExcelDocument(
                        id=row[0],
                        file_name=row[1],
                        file_size=row[2],
                        sheet_names=json.loads(row[3]),
                        metadata=json.loads(row[4]),
                        upload_time=datetime.fromisoformat(row[5]),
                        storage_type=row[6]
                    ))
                    
                return excel_files
                
        except sqlite3.Error as e:
            self.logger.error(f"Failed to list Excel files: {str(e)}")
            raise SQLiteDatabaseError(f"Excel file listing failed: {str(e)}")
            
    def delete_excel_file(self, file_id: str) -> bool:
        """
        Delete Excel file from database.
        
        Args:
            file_id: ID of the Excel file to delete
            
        Returns:
            True if deletion was successful, False otherwise
        """
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                cursor.execute("DELETE FROM excel_documents WHERE id = ?", (file_id,))
                conn.commit()
                
                deleted = cursor.rowcount > 0
                if deleted:
                    self.logger.info(f"Deleted Excel file with ID: {file_id}")
                else:
                    self.logger.warning(f"No Excel file found with ID: {file_id}")
                    
                return deleted
                
        except sqlite3.Error as e:
            self.logger.error(f"Failed to delete Excel file: {str(e)}")
            raise SQLiteDatabaseError(f"Excel file deletion failed: {str(e)}")
            
    def get_sheet_data(self, file_id: str, sheet_name: Optional[str] = None) -> List[ExcelSheetData]:
        """
        Get sheet data for an Excel file.
        
        Args:
            file_id: ID of the Excel file
            sheet_name: Optional specific sheet name
            
        Returns:
            List of ExcelSheetData objects
        """
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                if sheet_name:
                    cursor.execute("""
                        SELECT file_id, sheet_name, headers, row_count, column_count, sample_data, data_types
                        FROM excel_sheet_data 
                        WHERE file_id = ? AND sheet_name = ?
                    """, (file_id, sheet_name))
                else:
                    cursor.execute("""
                        SELECT file_id, sheet_name, headers, row_count, column_count, sample_data, data_types
                        FROM excel_sheet_data 
                        WHERE file_id = ?
                    """, (file_id,))
                
                sheet_data_list = []
                for row in cursor.fetchall():
                    sheet_data_list.append(ExcelSheetData(
                        file_id=row[0],
                        sheet_name=row[1],
                        headers=json.loads(row[2]),
                        row_count=row[3],
                        column_count=row[4],
                        sample_data=json.loads(row[5]),
                        data_types=json.loads(row[6])
                    ))
                    
                return sheet_data_list
                
        except sqlite3.Error as e:
            self.logger.error(f"Failed to get sheet data: {str(e)}")
            raise SQLiteDatabaseError(f"Sheet data retrieval failed: {str(e)}")
            
    def search_excel_data(self, query: str, file_id: Optional[str] = None, 
                         sheet_name: Optional[str] = None, column_name: Optional[str] = None,
                         case_sensitive: bool = False, max_results: int = 50) -> List[Dict[str, Any]]:
        """
        Search for data within Excel files.
        
        Args:
            query: Search query string
            file_id: Optional specific file to search within
            sheet_name: Optional specific sheet to search within
            column_name: Optional specific column to search within
            case_sensitive: Whether search should be case sensitive
            max_results: Maximum number of results to return
            
        Returns:
            List of search results with file and location information
        """
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # Build the search query
                base_query = """
                    SELECT ed.file_name, esd.sheet_name, esd.sample_data
                    FROM excel_sheet_data esd
                    JOIN excel_documents ed ON esd.file_id = ed.id
                    WHERE esd.sample_data LIKE ?
                """
                
                params = [f"%{query}%"]
                
                # Add filters
                if file_id:
                    base_query += " AND esd.file_id = ?"
                    params.append(file_id)
                    
                if sheet_name:
                    base_query += " AND esd.sheet_name = ?"
                    params.append(sheet_name)
                
                # Add limit
                base_query += " LIMIT ?"
                params.append(max_results)
                
                cursor.execute(base_query, params)
                
                results = []
                for row in cursor.fetchall():
                    file_name, sheet_name, sample_data_json = row
                    sample_data = json.loads(sample_data_json)
                    
                    # Search within sample data
                    for row_data in sample_data:
                        for col_name, cell_value in row_data.items():
                            if column_name and col_name != column_name:
                                continue
                                
                            if cell_value and query.lower() in str(cell_value).lower():
                                results.append({
                                    'file_name': file_name,
                                    'sheet_name': sheet_name,
                                    'column_name': col_name,
                                    'cell_value': cell_value,
                                    'file_id': file_id
                                })
                                if len(results) >= max_results:
                                    return results
                
                return results
                
        except sqlite3.Error as e:
            self.logger.error(f"Failed to search Excel data: {str(e)}")
            raise SQLiteDatabaseError(f"Excel data search failed: {str(e)}")
            
    def get_database_info(self) -> Dict[str, Any]:
        """
        Get information about the SQLite database.
        
        Returns:
            Dictionary containing database statistics
        """
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # Get document count
                cursor.execute("SELECT COUNT(*) FROM excel_documents")
                doc_count = cursor.fetchone()[0]
                
                # Get sheet count
                cursor.execute("SELECT COUNT(*) FROM excel_sheet_data")
                sheet_count = cursor.fetchone()[0]
                
                # Get total file size
                cursor.execute("SELECT SUM(file_size) FROM excel_documents")
                total_size = cursor.fetchone()[0] or 0
                
                return {
                    'document_count': doc_count,
                    'sheet_count': sheet_count,
                    'total_size_bytes': total_size,
                    'database_path': self.db_path,
                    'database_size': os.path.getsize(self.db_path) if os.path.exists(self.db_path) else 0
                }
                
        except sqlite3.Error as e:
            self.logger.error(f"Failed to get database info: {str(e)}")
            return {'error': str(e)}
            
    def _create_dynamic_table_for_sheet(self, file_id: str, sheet_name: str, headers: List[str], 
                                      data_types: Dict[str, str], cursor: sqlite3.Cursor) -> str:
        """
        Create a dynamic table for an Excel sheet with proper column types.
        
        Args:
            file_id: ID of the Excel file
            sheet_name: Name of the sheet
            headers: List of column headers
            data_types: Dictionary mapping headers to SQLite data types
            cursor: Database cursor for executing queries
            
        Returns:
            Name of the created dynamic table
            
        Raises:
            SQLiteDatabaseError: If table creation fails
        """
        try:
            # Remove the 'excel_' prefix and sanitize the UUID part
            file_id_without_prefix = file_id.replace('excel_', '')
            sanitized_file_id = self._sanitize_for_table_name(file_id_without_prefix)
            
            # Generate table name from sanitized file_id and sheet_name
            table_name = f"excel_{sanitized_file_id}_{self._sanitize_for_table_name(sheet_name.lower())}"
            
            # Valid SQLite data types
            valid_sqlite_types = {"INTEGER", "REAL", "TEXT", "BLOB"}
            
            # Build CREATE TABLE SQL
            column_definitions = []
            for header in headers:
                python_type = data_types.get(header, "TEXT")
                # Ensure the type is a valid SQLite type
                if python_type in valid_sqlite_types:
                    sqlite_type = python_type
                else:
                    sqlite_type = "TEXT"
                
                # Sanitize column name
                column_name = self._sanitize_for_table_name(header)
                if not column_name:
                    column_name = f"column_{headers.index(header)}"
                    
                column_definitions.append(f'"{column_name}" {sqlite_type}')
            
            create_table_sql = f'CREATE TABLE IF NOT EXISTS "{table_name}" ({", ".join(column_definitions)})'
            
            # Create the table
            cursor.execute(create_table_sql)
            
            self.logger.info(f"Created dynamic table '{table_name}' for sheet '{sheet_name}'")
            return table_name
            
        except sqlite3.Error as e:
            self.logger.error(f"Failed to create dynamic table: {str(e)}")
            raise SQLiteDatabaseError(f"Dynamic table creation failed: {str(e)}")
            
    def _insert_sheet_data_into_dynamic_table(self, file_path: str, file_id: str, sheet_name: str, 
                                            table_name: str, cursor: sqlite3.Cursor) -> int:
        """
        Insert Excel sheet data into the dynamic table.
        
        Args:
            file_path: Path to the Excel file
            file_id: ID of the Excel file
            sheet_name: Name of the sheet
            table_name: Name of the dynamic table
            cursor: Database cursor for executing queries
            
        Returns:
            Number of rows inserted
            
        Raises:
            SQLiteDatabaseError: If data insertion fails
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook[sheet_name]
            
            # Extract headers (first row)
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value) if cell.value else f"Column_{cell.column}")
            
            # Sanitize headers for SQL
            sanitized_headers = []
            for header in headers:
                sanitized_header = "".join(c for c in header if c.isalnum() or c == '_')
                if not sanitized_header:
                    sanitized_header = f"column_{headers.index(header)}"
                sanitized_headers.append(sanitized_header)
            
            # Extract all data rows
            rows_inserted = 0
            for row_idx in range(2, sheet.max_row + 1):  # Start from row 2 (after header)
                row_data = []
                for col_idx, header in enumerate(headers, 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    
                    # Convert Python types to SQLite-compatible types
                    if cell_value is None:
                        row_data.append(None)
                    elif isinstance(cell_value, (int, float)):
                        row_data.append(cell_value)
                    elif isinstance(cell_value, bool):
                        row_data.append(1 if cell_value else 0)
                    elif isinstance(cell_value, datetime):
                        row_data.append(cell_value.isoformat())
                    else:
                        row_data.append(str(cell_value))
                
                # Build INSERT SQL
                placeholders = ", ".join(["?"] * len(sanitized_headers))
                insert_sql = f'INSERT INTO "{table_name}" ({", ".join([f"{h}" for h in sanitized_headers])}) VALUES ({placeholders})'
                
                try:
                    cursor.execute(insert_sql, row_data)
                    rows_inserted += 1
                except sqlite3.Error as e:
                    self.logger.warning(f"Failed to insert row {row_idx} into table {table_name}: {str(e)}")
                    continue
            
            self.logger.info(f"Inserted {rows_inserted} rows into dynamic table '{table_name}'")
            return rows_inserted
            
        except Exception as e:
            self.logger.error(f"Failed to insert data into dynamic table: {str(e)}")
            raise SQLiteDatabaseError(f"Data insertion failed: {str(e)}")
            
    def store_excel_file_with_dynamic_tables(self, file_path: str, file_name: str, file_size: int,
                                           sheet_names: List[str], metadata: Dict[str, Any]) -> Dict[str, Any]:
        """
        Store Excel file with dynamic table creation for each sheet.
        
        Args:
            file_path: Path to the Excel file
            file_name: Name of the Excel file
            file_size: Size of the file in bytes
            sheet_names: List of sheet names in the Excel file
            metadata: Additional metadata about the file
            
        Returns:
            Dictionary with file ID and table creation results
            
        Raises:
            SQLiteDatabaseError: If storage operation fails
        """
        try:
            # Create ExcelDocument object for validation
            excel_doc = ExcelDocument(
                file_name=file_name,
                file_size=file_size,
                sheet_names=sheet_names,
                metadata=metadata
            )
            
            table_results = {}
            
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # Insert Excel document
                cursor.execute("""
                    INSERT INTO excel_documents 
                    (id, file_name, file_size, sheet_names, metadata, upload_time)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (
                    excel_doc.id,
                    excel_doc.file_name,
                    excel_doc.file_size,
                    json.dumps(excel_doc.sheet_names),
                    json.dumps(excel_doc.metadata),
                    excel_doc.upload_time.isoformat()
                ))
                
                # Process each sheet and create dynamic tables
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    # Extract headers and infer data types
                    headers = []
                    data_types = {}
                    
                    for cell in sheet[1]:
                        headers.append(str(cell.value) if cell.value else f"Column_{cell.column}")
                    
                    # Infer data types from first few rows
                    for row_idx in range(2, min(7, sheet.max_row + 1)):
                        for col_idx, header in enumerate(headers, 1):
                            cell_value = sheet.cell(row=row_idx, column=col_idx).value
                            if header not in data_types:
                                data_types[header] = self._infer_data_type(cell_value)
                    
                    # Create dynamic table
                    table_name = self._create_dynamic_table_for_sheet(
                        excel_doc.id, sheet_name, headers, data_types, cursor
                    )
                    
                    # Insert data into dynamic table
                    rows_inserted = self._insert_sheet_data_into_dynamic_table(
                        file_path, excel_doc.id, sheet_name, table_name, cursor
                    )
                    
                    # Store sample data for search functionality
                    sample_data = []
                    for row_idx in range(2, min(7, sheet.max_row + 1)):
                        row_data = {}
                        for col_idx, header in enumerate(headers, 1):
                            cell_value = sheet.cell(row=row_idx, column=col_idx).value
                            row_data[header] = cell_value
                        sample_data.append(row_data)
                    
                    # Create ExcelSheetData object
                    sheet_data = ExcelSheetData(
                        file_id=excel_doc.id,
                        sheet_name=sheet_name,
                        headers=headers,
                        row_count=sheet.max_row - 1,
                        column_count=sheet.max_column,
                        sample_data=sample_data,
                        data_types=data_types
                    )
                    
                    # Insert sheet metadata
                    cursor.execute("""
                        INSERT INTO excel_sheet_data 
                        (file_id, sheet_name, headers, row_count, column_count, sample_data, data_types)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (
                        sheet_data.file_id,
                        sheet_data.sheet_name,
                        json.dumps(sheet_data.headers),
                        sheet_data.row_count,
                        sheet_data.column_count,
                        json.dumps(sheet_data.sample_data),
                        json.dumps(sheet_data.data_types)
                    ))
                    
                    table_results[sheet_name] = {
                        'table_name': table_name,
                        'rows_inserted': rows_inserted,
                        'headers': headers,
                        'data_types': data_types
                    }
                
                conn.commit()
                
                result = {
                    'file_id': excel_doc.id,
                    'file_name': file_name,
                    'tables_created': table_results,
                    'total_sheets': len(sheet_names),
                    'status': 'success'
                }
                
                self.logger.info(f"Successfully stored Excel file with dynamic tables: {file_name}")
                return result
                
        except sqlite3.Error as e:
            self.logger.error(f"Failed to store Excel file with dynamic tables: {str(e)}")
            raise SQLiteDatabaseError(f"Excel file storage with dynamic tables failed: {str(e)}")
        except Exception as e:
            self.logger.error(f"Unexpected error storing Excel file: {str(e)}")
            raise SQLiteDatabaseError(f"Unexpected error: {str(e)}")
            
    def execute_sql_query(self, sql_query: str) -> List[Dict[str, Any]]:
        """
        Execute a SQL query on the SQLite database.
        
        Args:
            sql_query: SQL query to execute
            
        Returns:
            List of dictionaries representing query results
            
        Raises:
            SQLiteDatabaseError: If query execution fails
        """
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row  # This enables column access by name
                cursor = conn.cursor()
                
                cursor.execute(sql_query)
                
                results = []
                if sql_query.strip().upper().startswith('SELECT'):
                    # For SELECT queries, return the results
                    rows = cursor.fetchall()
                    for row in rows:
                        results.append(dict(row))
                else:
                    # For non-SELECT queries, return execution info
                    results.append({
                        'rows_affected': cursor.rowcount,
                        'query_type': 'non_select',
                        'message': 'Query executed successfully'
                    })
                    conn.commit()
                
                return results
                
        except sqlite3.Error as e:
            self.logger.error(f"SQL query execution failed: {str(e)}")
            raise SQLiteDatabaseError(f"SQL query execution failed: {str(e)}")
            
    def get_dynamic_table_schemas(self, file_id: Optional[str] = None) -> Dict[str, Any]:
        """
        Get schema information for all dynamic tables.
        
        Args:
            file_id: Optional specific file ID to get schemas for
            
        Returns:
            Dictionary containing table schema information
        """
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # Get all tables in the database
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                all_tables = [row[0] for row in cursor.fetchall()]
                
                # Filter for dynamic tables (those starting with 'excel_')
                dynamic_tables = [table for table in all_tables if table.startswith('excel_')]
                
                if file_id:
                    # Filter tables for specific file using sanitized file ID
                    # Remove 'excel_' prefix and sanitize for comparison
                    file_id_without_prefix = file_id.replace('excel_', '')
                    sanitized_file_id = self._sanitize_for_table_name(file_id_without_prefix)
                    file_prefix = f"excel_{sanitized_file_id}_"
                    dynamic_tables = [table for table in dynamic_tables if table.startswith(file_prefix)]
                
                schemas = {}
                for table_name in dynamic_tables:
                    cursor.execute(f"PRAGMA table_info(\"{table_name}\")")
                    columns = cursor.fetchall()
                    
                    column_info = []
                    for col in columns:
                        column_info.append({
                            'name': col[1],
                            'type': col[2],
                            'not_null': bool(col[3]),
                            'default_value': col[4],
                            'primary_key': bool(col[5])
                        })
                    
                    # Get row count
                    cursor.execute(f'SELECT COUNT(*) FROM "{table_name}"')
                    row_count = cursor.fetchone()[0]
                    
                    schemas[table_name] = {
                        'columns': column_info,
                        'row_count': row_count,
                        'file_id': self._extract_file_id_from_table_name(table_name)
                    }
                
                return schemas
                
        except sqlite3.Error as e:
            self.logger.error(f"Failed to get dynamic table schemas: {str(e)}")
            raise SQLiteDatabaseError(f"Failed to get dynamic table schemas: {str(e)}")
            
    def _extract_file_id_from_table_name(self, table_name: str) -> Optional[str]:
        """
        Extract file ID from dynamic table name.
        
        Args:
            table_name: Name of the dynamic table
            
        Returns:
            File ID if found, None otherwise
        """
        try:
            # Table name format: excel_{sanitized_file_id}_{sheet_name}
            # Example: "excel_77e71b80a8dd450f80bee61ba1b1482d_sheet1"
            parts = table_name.split('_')
            if len(parts) >= 2:
                # parts[1] is the sanitized UUID without hyphens
                sanitized_uuid = parts[1]
                # Reconstruct the original file ID with hyphens
                return self._reconstruct_file_id(sanitized_uuid)
            return None
        except Exception:
            return None
            
    def clear_database(self) -> Dict[str, Any]:
        """
        Clear all data from the SQLite database.
        
        Drops all tables and recreates the base tables.
        
        Returns:
            Dictionary with operation results including success flag and tables dropped count.
        """
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # Disable foreign key constraints to allow dropping tables in any order
                cursor.execute("PRAGMA foreign_keys = OFF")
                
                # Get all table names
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                tables = [row[0] for row in cursor.fetchall()]
                
                # System tables that should not be dropped
                system_tables = {'sqlite_sequence', 'sqlite_stat1', 'sqlite_stat2', 
                                 'sqlite_stat3', 'sqlite_stat4', 'sqlite_master'}
                # Also exclude any table starting with 'sqlite_'
                user_tables = [t for t in tables if not t.startswith('sqlite_') and t not in system_tables]
                
                # Drop all user tables
                tables_dropped = 0
                for table_name in user_tables:
                    try:
                        cursor.execute(f'DROP TABLE IF EXISTS "{table_name}"')
                        tables_dropped += 1
                    except sqlite3.Error as e:
                        self.logger.error(f"Failed to drop table {table_name}: {str(e)}")
                
                # Re-enable foreign key constraints
                cursor.execute("PRAGMA foreign_keys = ON")
                
                # Clear data from system tables (sqlite_sequence) if it exists
                if 'sqlite_sequence' in tables:
                    try:
                        cursor.execute("DELETE FROM sqlite_sequence")
                    except sqlite3.Error as e:
                        self.logger.warning(f"Failed to clear sqlite_sequence: {str(e)}")
                
                conn.commit()
                
            # Reinitialize the database (creates base tables if they don't exist)
            self._init_database()
            
            return {
                "success": True,
                "tables_dropped": tables_dropped,
                "message": f"Successfully dropped {tables_dropped} tables and reinitialized database"
            }
            
        except sqlite3.Error as e:
            self.logger.error(f"Failed to clear database: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "message": f"Failed to clear database: {str(e)}"
            }

    def health_check(self) -> bool:
        """
        Perform a health check on the SQLite database.

        Returns:
            True if database is healthy, False otherwise
        """
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT 1")
                return True
        except sqlite3.Error:
            return False
